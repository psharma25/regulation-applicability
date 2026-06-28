"""Generate the downloadable Excel workbooks (risk assessment + security requirements)."""
import os, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, "..", "backend", "data")
def load(n): return json.load(open(os.path.join(DATA, n + ".json")))

HEAD = PatternFill("solid", fgColor="5F7E9E"); HEADF = Font(color="FFFFFF", bold=True, size=10)
TITLE = Font(bold=True, size=14, color="33384A"); SUB = Font(size=9, color="566377", italic=True)
WRAP = Alignment(wrap_text=True, vertical="top"); THIN = Side(style="thin", color="E1E7F0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALT = PatternFill("solid", fgColor="F6F9FC")

def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c); cell.fill = HEAD; cell.font = HEADF
        cell.alignment = Alignment(wrap_text=True, vertical="center"); cell.border = BORDER

def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, 1): ws.column_dimensions[get_column_letter(i)].width = w

# ---------------- Risk Assessment workbook ----------------
def risk_wb():
    from openpyxl.formatting.rule import CellIsRule
    wb = Workbook(); ws = wb.active; ws.title = "Risk Register"
    ws["A1"] = "Medical Device Cybersecurity Risk Assessment"; ws["A1"].font = TITLE
    ws["A2"] = "Cyber issue → C/I/A impact → inherent risk → mitigating control → hazardous situation → patient-safety impact → residual risk → recommendation → standards. Severity & Likelihood 1–5; Risk = Sev×Lik (live formulas). Illustrative — adapt to your device."; ws["A2"].font = SUB
    hdr = ["ID", "Cyber issue", "C/I/A impact", "Inh. Sev", "Inh. Lik", "Inherent risk", "Inherent level",
           "Mitigating control", "Hazardous situation", "Patient-safety impact",
           "Res. Sev", "Res. Lik", "Residual risk", "Residual level", "Recommendation to remediate", "Standards"]
    r0 = 4
    for i, h in enumerate(hdr, 1): ws.cell(row=r0, column=i, value=h)
    style_header(ws, r0, len(hdr))
    rows = load("risk")
    for j, row in enumerate(rows):
        rr = r0 + 1 + j
        vals = [row["id"], row["cyber_issue"], row["cia"], row["inh_sev"], row["inh_lik"], None, None,
                row["control"], row["hazard"], row["patient"], row["res_sev"], row["res_lik"], None, None,
                row["rec"], row["std"]]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=rr, column=i, value=v); c.alignment = WRAP; c.border = BORDER
            if j % 2: c.fill = ALT
        # live formulas: inherent risk (F) = Sev*Lik, level (G); residual risk (M) = Sev*Lik, level (N)
        ws.cell(row=rr, column=6, value=f"=D{rr}*E{rr}")
        ws.cell(row=rr, column=7, value=f'=IF(F{rr}>=12,"High",IF(F{rr}>=6,"Medium","Low"))')
        ws.cell(row=rr, column=13, value=f"=K{rr}*L{rr}")
        ws.cell(row=rr, column=14, value=f'=IF(M{rr}>=12,"High",IF(M{rr}>=6,"Medium","Low"))')
        for col in (4, 5, 6, 7, 11, 12, 13, 14):
            ws.cell(row=rr, column=col).alignment = Alignment(horizontal="center", vertical="top")
    last = r0 + len(rows)
    # conditional formatting on the computed risk numbers
    for colL in ("F", "M"):
        rng = f"{colL}{r0+1}:{colL}{last}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual", formula=["12"], fill=PatternFill("solid", fgColor="F2AEBE")))
        ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["6", "11"], fill=PatternFill("solid", fgColor="EAC873")))
        ws.conditional_formatting.add(rng, CellIsRule(operator="lessThanOrEqual", formula=["5"], fill=PatternFill("solid", fgColor="B6D4A8")))
    widths(ws, [7, 30, 12, 8, 8, 10, 11, 30, 26, 28, 8, 8, 10, 11, 34, 22])
    ws.freeze_panes = "B5"
    # scoring + references sheets
    s2 = wb.create_sheet("Scoring key"); s2["A1"] = "Scoring key"; s2["A1"].font = TITLE
    for i, h in enumerate(["Scale", "1", "2", "3", "4", "5"], 1): s2.cell(row=3, column=i, value=h)
    style_header(s2, 3, 6)
    s2.append(["Severity", "Negligible", "Minor", "Serious", "Critical", "Catastrophic"])
    s2.append(["Likelihood", "Improbable", "Remote", "Occasional", "Probable", "Frequent"])
    s2.append(["Risk = Sev×Lik", "1-5 Low", "6-11 Medium", "12-25 High", "", ""])
    s2.append(["C/I/A", "Confidentiality", "Integrity", "Availability", "", ""])
    widths(s2, [16, 22, 22, 22, 16, 16])
    s3 = wb.create_sheet("References"); s3["A1"] = "References"; s3["A1"].font = TITLE
    refs = [("ISO 14971:2019", "Application of risk management to medical devices"),
            ("ISO/TR 24971:2020", "Guidance on the application of ISO 14971"),
            ("AAMI TIR57:2016", "Principles for medical device security risk management"),
            ("ANSI/AAMI SW96:2023", "Security risk management for device manufacturers"),
            ("FDA (2025)", "Cybersecurity in Medical Devices — premarket guidance"),
            ("FDA (2016)", "Postmarket Management of Cybersecurity in Medical Devices"),
            ("IEC 81001-5-1:2021", "Health software security activities in the lifecycle"),
            ("NIST SP 800-30 Rev.1", "Guide for conducting risk assessments"),
            ("NIST AI RMF 1.0", "AI risk management (for AI-enabled devices)")]
    s3.cell(row=3, column=1, value="Standard"); s3.cell(row=3, column=2, value="Title"); style_header(s3, 3, 2)
    for i, (a, b) in enumerate(refs): s3.cell(row=4 + i, column=1, value=a); s3.cell(row=4 + i, column=2, value=b)
    widths(s3, [26, 60])
    out = os.path.join(HERE, "risk-assessment-medical-device.xlsx"); wb.save(out); print("wrote", os.path.basename(out), "(with formulas)")

# ---------------- Security Requirements workbook ----------------
def secreq_wb():
    wb = Workbook(); ws = wb.active; ws.title = "Security Requirements"
    ws["A1"] = "Security Requirements & V&V Tests"; ws["A1"].font = TITLE
    ws["A2"] = "Per-requirement validation mapping. Filter the Profiles column for AI / Cloud / Mobile / Firmware, etc. Illustrative — adapt to your device."; ws["A2"].font = SUB
    hdr = ["ID", "Security requirement", "Description", "Automatable / Manual", "V&V test to validate", "Applicable profiles"]
    r0 = 4
    for i, h in enumerate(hdr, 1): ws.cell(row=r0, column=i, value=h)
    style_header(ws, r0, len(hdr))
    for j, row in enumerate(load("secreqs")):
        rr = r0 + 1 + j
        vals = [row["id"], row["req"], row["desc"], row["mode"], row["vv"], ", ".join(row["tech"])]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=rr, column=i, value=v); c.alignment = WRAP; c.border = BORDER
            if j % 2: c.fill = ALT
    widths(ws, [12, 28, 40, 18, 40, 34]); ws.freeze_panes = "A5"; ws.auto_filter.ref = f"A4:F{r0 + len(load('secreqs'))}"
    out = os.path.join(HERE, "security-requirements-vv.xlsx"); wb.save(out); print("wrote", os.path.basename(out))

risk_wb(); secreq_wb()
