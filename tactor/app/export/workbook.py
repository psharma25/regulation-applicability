"""Excel scoring workbook generator.

Produces a multi-sheet .xlsx:
  - Scorecard:  one row per actor, composites + each criterion, color-graded
  - Rubric:     the criteria, weights, and how raw values map to composites
  - Evidence:   the fresh sources gathered this scan, per actor
  - About:      methodology + honest data caveats
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..config import load_config
from ..models import ScanResult
from ..scoring.rubric import reliability_band, severity_band

INK = "0E1420"
AMBER = "E8A33D"
TEAL = "3DA89A"
RED = "C7544A"
PARCH = "E6E1D3"

HEADER_FILL = PatternFill("solid", fgColor=INK)
HEADER_FONT = Font(color=PARCH, bold=True, size=11)
TITLE_FONT = Font(color=INK, bold=True, size=14)
THIN = Side(style="thin", color="C9C4B5")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _grade_fill(score: float, reverse: bool = False) -> PatternFill:
    """Amber gradient for severity; teal gradient for reliability (reverse)."""
    t = (score - 1) / 9.0
    if reverse:  # reliability: low score = red, high = teal
        if score >= 7:
            return PatternFill("solid", fgColor="CBE7E1")
        if score >= 5:
            return PatternFill("solid", fgColor="F3E6C8")
        return PatternFill("solid", fgColor="F1D2CD")
    # severity: high = deep amber/red
    if score >= 8.5:
        return PatternFill("solid", fgColor="EBB9A9")
    if score >= 7:
        return PatternFill("solid", fgColor="F4D9B0")
    if score >= 5:
        return PatternFill("solid", fgColor="F6EBCF")
    return PatternFill("solid", fgColor="E6EFE0")


def _style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def build_workbook(result: ScanResult, out_dir: Path) -> Path:
    crit_cfg = load_config()["rubric"]["criteria"]
    crit_keys = list(crit_cfg.keys())

    wb = Workbook()

    # --- Scorecard ---------------------------------------------------------
    ws = wb.active
    ws.title = "Scorecard"
    ws["A1"] = "Threat Actor Reputation Scorecard"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        f"Scan {result.scan_id} · {result.finished or result.started} UTC · "
        f"backend: {result.llm_backend} · "
        f"{len(result.actors)} actors ({result.discovered_count} discovered this scan)"
    )
    ws["A2"].font = Font(italic=True, color="555049", size=9)

    headers = (
        ["Actor", "Source", "Status", "First seen", "Origin", "Model", "Sectors",
         "Reported victims", "Nation-state", "Recent victims",
         "Threat severity", "Severity band", "Deal reliability",
         "Reliability band", "Confidence"]
        + [crit_cfg[k]["label"] for k in crit_keys]
        + ["Notes"]
    )
    hrow = 4
    for i, h in enumerate(headers, start=1):
        ws.cell(row=hrow, column=i, value=h)
    _style_header(ws, hrow, len(headers))

    r = hrow + 1
    for a in result.actors:
        col = 1
        ws.cell(row=r, column=col, value=a.name).font = Font(bold=True); col += 1
        src = ws.cell(row=r, column=col, value="discovered" if a.discovered else "seed")
        if a.discovered:
            src.fill = PatternFill("solid", fgColor="DEF0EC")
            src.font = Font(bold=True, color="2C7D71")
        col += 1
        st = ws.cell(row=r, column=col, value=a.status)
        if a.status == "defunct":
            st.font = Font(italic=True, color="7A8290")
        col += 1
        ws.cell(row=r, column=col, value=a.first_seen); col += 1
        ws.cell(row=r, column=col, value=a.origin); col += 1
        ws.cell(row=r, column=col, value=a.model); col += 1
        ws.cell(row=r, column=col, value=", ".join(a.sectors)); col += 1
        ws.cell(row=r, column=col, value=", ".join(a.notable_victims)); col += 1
        ns = ws.cell(row=r, column=col, value=a.nation_state_class)
        if a.nation_state_class == "directed":
            ns.fill = PatternFill("solid", fgColor="F1D2CD")
            ns.font = Font(bold=True, color=RED)
        col += 1
        ws.cell(row=r, column=col, value=a.recent_victims); col += 1

        sev = ws.cell(row=r, column=col, value=a.threat_severity)
        sev.fill = _grade_fill(a.threat_severity); col += 1
        ws.cell(row=r, column=col, value=severity_band(a.threat_severity)); col += 1

        rel = ws.cell(row=r, column=col, value=a.deal_reliability)
        rel.fill = _grade_fill(a.deal_reliability, reverse=True); col += 1
        ws.cell(row=r, column=col, value=reliability_band(a.deal_reliability)); col += 1
        ws.cell(row=r, column=col, value=a.confidence); col += 1

        for k in crit_keys:
            cs = a.criteria.get(k)
            cell = ws.cell(row=r, column=col, value=cs.raw if cs else None)
            if cs:
                rev = crit_cfg[k].get("higher_is_worse", True) is False
                cell.fill = _grade_fill(cs.raw, reverse=rev)
                if cs.rationale:
                    cell.comment = _comment(cs.rationale, cs.source)
            col += 1
        ws.cell(row=r, column=col, value=a.notes)
        r += 1

    _autosize(ws, headers)
    ws.freeze_panes = "B5"

    # --- Rubric ------------------------------------------------------------
    rb = wb.create_sheet("Rubric")
    rb["A1"] = "Scoring rubric"; rb["A1"].font = TITLE_FONT
    rhead = ["Criterion", "Description", "Severity weight", "Reliability weight",
             "Higher value means"]
    for i, h in enumerate(rhead, start=1):
        rb.cell(row=3, column=i, value=h)
    _style_header(rb, 3, len(rhead))
    rr = 4
    for k, m in crit_cfg.items():
        rb.cell(row=rr, column=1, value=m["label"])
        rb.cell(row=rr, column=2, value=m["description"])
        rb.cell(row=rr, column=3, value=m.get("weight_severity", 0))
        rb.cell(row=rr, column=4, value=m.get("weight_reliability", 0))
        rb.cell(row=rr, column=5,
                value="more severe" if m.get("higher_is_worse", True)
                else "more reliable (honors deals)")
        rr += 1
    _autosize(rb, rhead)

    # --- Evidence ----------------------------------------------------------
    ev = wb.create_sheet("Evidence")
    ev["A1"] = "Evidence gathered this scan"; ev["A1"].font = TITLE_FONT
    ehead = ["Actor", "Source", "Title", "Snippet", "URL", "Published"]
    for i, h in enumerate(ehead, start=1):
        ev.cell(row=3, column=i, value=h)
    _style_header(ev, 3, len(ehead))
    er = 4
    for a in result.actors:
        for e in a.evidence:
            ev.cell(row=er, column=1, value=a.name)
            ev.cell(row=er, column=2, value=e.source)
            ev.cell(row=er, column=3, value=e.title)
            ev.cell(row=er, column=4, value=e.snippet)
            ev.cell(row=er, column=5, value=e.url)
            ev.cell(row=er, column=6, value=e.published)
            er += 1
    if er == 4:
        ev.cell(row=4, column=1,
                value="No fresh evidence gathered (sources unreachable or "
                      "offline). Scores reflect analyst priors.")
    _autosize(ev, ehead, cap=70)

    # --- About -------------------------------------------------------------
    ab = wb.create_sheet("About")
    ab["A1"] = "Methodology & data caveats"; ab["A1"].font = TITLE_FONT
    lines = [
        "",
        "Two composite indices are produced per actor:",
        "  Threat Severity (1-10)  — how dangerous / active / capable. Higher = worse.",
        "  Deal Reliability (1-10) — if forced to pay, how likely the actor honors it.",
        "      Higher = more 'reliable' criminal. This is a planning input, NOT an",
        "      endorsement; all listed actors are criminal operations.",
        "",
        "DATA CAVEATS — read before acting on any number:",
        "  1. Ransoms CHARGED are sometimes visible (leak-site demands, leaked chats).",
        "     Ransoms PAID and the identities of payers are mostly undisclosed.",
        "     Dollar-scale scores are directional, not audited ledgers.",
        "  2. Per-group decryptor success rates are tracked privately by negotiation",
        "     firms and rarely published. Reliability scores reflect reputation and",
        "     incentive structure (durable RaaS brands have reason to honor deals;",
        "     transient/collapsing groups do not), not a verified success percentage.",
        "  3. Across all groups, ~1/3 of victims who pay receive non-functional or",
        "     incomplete decryptors. Payment never guarantees recovery or deletion",
        "     of exfiltrated data.",
        "  4. Paying a state-directed actor (e.g. DPRK/Lazarus) may itself violate",
        "     OFAC sanctions. Treat 'directed' nation-state class as a legal red flag.",
        "",
        "Scores combine analyst priors with fresh public evidence gathered at scan",
        "time (ransomware.live, CISA advisories, configurable threat-intel RSS).",
        "Re-run the scan to refresh.",
    ]
    for i, ln in enumerate(lines, start=2):
        ab.cell(row=i, column=1, value=ln)
    ab.column_dimensions["A"].width = 80

    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    path = out_dir / f"threat_actor_scorecard_{stamp}.xlsx"
    wb.save(path)
    return path


def _comment(text: str, source: str):
    from openpyxl.comments import Comment
    return Comment(f"{text}\n— {source}", "tracker")


def _autosize(ws, headers, cap: int = 40) -> None:
    for i, h in enumerate(headers, start=1):
        letter = get_column_letter(i)
        longest = len(str(h))
        for cell in ws[letter]:
            if cell.value:
                longest = max(longest, min(len(str(cell.value)), cap))
        ws.column_dimensions[letter].width = min(longest + 2, cap)
