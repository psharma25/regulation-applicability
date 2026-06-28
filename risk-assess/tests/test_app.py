import os
import tempfile

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.main import app
from backend import excel

client = TestClient(app)


def test_health():
    r = client.get("/api/health")
    assert r.status_code == 200
    assert r.json()["status"] == "ok"


def test_domains():
    r = client.get("/api/domains")
    ids = {d["id"] for d in r.json()["domains"]}
    assert ids == {"medical", "finance", "technology"}


def test_chat_qa():
    r = client.post("/api/chat", json={"message": "How does ISO 14971 estimate probability of harm?"})
    assert r.status_code == 200
    assert r.json()["action"] == "answer"


def test_chat_generate_routes_to_download():
    r = client.post("/api/chat", json={"message": "download the finance template"})
    d = r.json()
    assert d["action"] == "download" and d["domain"] == "finance"


def test_library_each_domain():
    expected = {"medical": "Hazard Mapping", "finance": "Risk Taxonomy", "technology": "Threat Catalog"}
    for dom, sheet in expected.items():
        r = client.get(f"/api/library/{dom}")
        assert r.status_code == 200
        d = r.json()
        assert d["sheet"] == sheet
        assert len(d["rows"]) >= 8 and len(d["columns"]) >= 5


def test_workbooks_have_mapping_sheet():
    from openpyxl import load_workbook
    expected = {"medical": "Hazard Mapping", "finance": "Risk Taxonomy", "technology": "Threat Catalog"}
    for dom, sheet in expected.items():
        p = os.path.join(tempfile.gettempdir(), f"_map_{dom}.xlsx")
        excel.generate(dom, p)
        assert sheet in load_workbook(p).sheetnames


def test_download_each_domain():
    for dom in ("medical", "finance", "technology"):
        r = client.get(f"/api/forms/{dom}/download")
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers["content-type"]


def test_generated_workbooks_have_formulas():
    for dom in ("medical", "finance", "technology"):
        p = os.path.join(tempfile.gettempdir(), f"_test_{dom}.xlsx")
        excel.generate(dom, p)
        wb = load_workbook(p)
        ws = wb.worksheets[0]
        formulas = [c.value for row in ws.iter_rows() for c in row
                    if isinstance(c.value, str) and c.value.startswith("=")]
        assert len(formulas) > 5
