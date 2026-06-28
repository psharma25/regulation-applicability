"""Medical device risk assessment workbook — ISO 14971:2019, with hazard mapping."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from . import common as C
from .library import LIBRARIES

DOMAIN = "medical"
TITLE = "Medical Device Risk Assessment (ISO 14971:2019)"
LIB = LIBRARIES["medical"]


def _scales(wb):
    ws = wb.create_sheet("Scales")
    ws.sheet_view.showGridLines = False
    C.set_widths(ws, {"A": 3, "B": 9, "C": 24, "D": 50, "F": 10, "G": 22, "H": 40, "O": 16, "P": 6})
    C.title(ws["B2"], "Scoring Scales & Risk Acceptability", 14)

    ws["B4"] = "SEVERITY OF HARM (S)"; C.style_header(ws["B4"]); ws.merge_cells("B4:D4")
    for i, h in enumerate(["Score", "Category", "Definition (patient / user harm)"]):
        C.style_header(ws.cell(5, 2 + i), fill=C.NAVY)
    sev = [(1, "Negligible", "No injury or slight inconvenience; no intervention."),
           (2, "Minor", "Temporary injury not requiring professional intervention."),
           (3, "Serious", "Injury requiring professional medical intervention."),
           (4, "Critical", "Permanent impairment or reversible life-threatening injury."),
           (5, "Catastrophic", "Death or permanent loss of life-sustaining function.")]
    for j, (s, cat, d) in enumerate(sev):
        C.text_cell(ws.cell(6 + j, 2), s, center=True); C.text_cell(ws.cell(6 + j, 3), cat); C.text_cell(ws.cell(6 + j, 4), d)

    ws["B12"] = "PROBABILITY OF HARM (P, via P1 x P2)"; C.style_header(ws["B12"]); ws.merge_cells("B12:D12")
    for i, h in enumerate(["Score", "Qualitative", "Indicative frequency"]):
        C.style_header(ws.cell(13, 2 + i), fill=C.NAVY)
    prob = [(1, "Improbable", "< 1 in 1,000,000"), (2, "Remote", "1 in 100k-1M"),
            (3, "Occasional", "1 in 1k-100k"), (4, "Probable", "1 in 100-1k"), (5, "Frequent", "> 1 in 100")]
    for j, (s, q, f) in enumerate(prob):
        C.text_cell(ws.cell(14 + j, 2), s, center=True); C.text_cell(ws.cell(14 + j, 3), q); C.text_cell(ws.cell(14 + j, 4), f)
    ws.cell(20, 2, "P1 = prob. hazardous situation occurs; P2 = prob. it leads to harm. P = ROUNDUP(P1*P2/5,0).")\
        .font = Font(name=C.ARIAL, italic=True, size=9, color="595959")
    ws.merge_cells("B20:D20")

    thr = C.add_threshold_block(ws, 4, anchor_col=6, score_label="ACCEPTABILITY (Risk Index = S x P, 1-25)")
    C.add_matrix(ws, 12, anchor_col=6, rows_label="Severity", cols_label="Probability")
    return thr


HEADERS = [
    ("Risk ID", 9), ("Feature / Item / Function", 20), ("Hazard category", 16), ("Hazard", 20),
    ("Foreseeable sequence of events (cause)", 26), ("Hazardous situation", 22), ("Harm", 20), ("Affected party", 13),
    ("S", 5), ("P1", 5), ("P2", 5), ("P", 6), ("Risk Index", 9), ("Risk Level", 12), ("Acceptability", 14),
    ("Risk control measure(s)", 26), ("Control type", 16), ("Control verification (effectiveness)", 24),
    ("V&V / Traceability ref.", 16),
    ("S'", 5), ("P1'", 5), ("P2'", 5), ("P'", 6), ("Residual Index", 9), ("Residual Level", 12), ("Residual Acceptability", 14),
    ("New risk introduced?", 11), ("Benefit-risk justification", 24), ("Residual risk disclosed", 14), ("Status", 10),
]
# text(1-7 except 3,8 inputs) | inputs C,H,I,J,K,S',P1',P2', etc.
EXAMPLES = [
    ["RA-001", "Infusion pump dosing engine", "Software fault", "Dose miscalculation after unit change",
     "SW miscomputes rate after unit-of-measure change", "Infusion faster than prescribed", "Overdose - serious injury", "Patient",
     4, 3, 4, "Dose-rate hard limit + independent watchdog", "Protective measure",
     "Fault-injection bench test; watchdog halts within 1 s", "SW-VER-218; SRS-4.2", 4, 1, 2, "No", "N/A - residual acceptable", "Yes - IFU Warnings"],
    ["RA-002", "Connected device remote interface", "Cybersecurity", "Unauthorized access (STRIDE: Tampering)",
     "Weak auth lets attacker alter therapy params", "Malicious change to therapy settings", "Inappropriate therapy - critical injury", "Patient",
     4, 3, 4, "Mutual TLS + signed updates + input validation", "Inherently safe design",
     "Pen-test confirms auth-bypass closed", "THREAT-MODEL-7; SBOM-3.1", 4, 1, 2, "No", "N/A - residual acceptable", "Yes - Security manual"],
    ["RA-003", "Reusable surgical tool", "Biological (infection)", "Inadequate reprocessing",
     "IFU reprocessing steps unclear to user", "Residual bioburden on device", "Infection - serious injury", "Patient",
     3, 2, 3, "Validated reprocessing IFU + user training", "Information for safety",
     "Reprocessing validation per ISO 17664", "REP-VAL-09", 3, 1, 2, "No", "N/A - residual acceptable", "Yes - IFU"],
    ["RA-004", "Patient monitor alarm subsystem", "Functional / loss of performance", "Missed critical alarm",
     "Alarm thread starves under high CPU load", "Clinician not alerted to deterioration", "Delayed treatment - serious injury", "Patient",
     4, 2, 3, "Watchdog reset + independent audible backup alarm", "Protective measure",
     "Load test verifies alarm under max load", "SW-VER-104", 4, 1, 2, "No", "N/A - residual acceptable", "Yes - IFU"],
]


def build(path):
    wb = Workbook()
    wb.remove(wb.active)
    t_int, t_high, t_med = _scales(wb)

    ws = wb.create_sheet("Risk Assessment", 0)
    ws.sheet_view.showGridLines = False
    ws.cell(1, 1, TITLE).font = Font(name=C.ARIAL, bold=True, color=C.NAVY, size=13)
    ws.row_dimensions[2].height = 44
    for i, (h, w) in enumerate(HEADERS, start=1):
        C.style_header(ws.cell(2, i, h))
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C3"

    def fill_formulas(r):
        ws.cell(r, 12, f'=IF(OR(J{r}="",K{r}=""),"",ROUNDUP(J{r}*K{r}/5,0))')   # P
        ws.cell(r, 13, f'=IF(OR(I{r}="",L{r}=""),"",I{r}*L{r})')                 # Risk Index
        ws.cell(r, 14, C.level_formula(f"M{r}", t_int, t_high, t_med))
        ws.cell(r, 15, C.accept_formula(f"M{r}", t_int, t_high, t_med))
        ws.cell(r, 23, f'=IF(OR(U{r}="",V{r}=""),"",ROUNDUP(U{r}*V{r}/5,0))')    # P'
        ws.cell(r, 24, f'=IF(OR(T{r}="",W{r}=""),"",T{r}*W{r})')                 # Residual Index
        ws.cell(r, 25, C.level_formula(f"X{r}", t_int, t_high, t_med))
        ws.cell(r, 26, C.accept_formula(f"X{r}", t_int, t_high, t_med))
        ws.cell(r, 30, f'=IF(Z{r}="","",IF(OR(Z{r}="Unacceptable",O{r}="Unacceptable"),"OPEN","Closed"))')
        for col in (12, 13, 14, 15, 23, 24, 25, 26, 30):
            ws.cell(r, col).alignment = Alignment(horizontal="center", wrap_text=True)
            ws.cell(r, col).border = C.BORDER

    start = 3
    text_cols = {1: 0, 2: 1, 4: 3, 5: 4, 6: 5, 7: 6, 16: 11, 18: 13, 19: 14, 28: 18}
    input_cols = {3: 2, 8: 7, 9: 8, 10: 9, 11: 10, 17: 12, 20: 15, 21: 16, 22: 17, 27: 19, 29: 20}
    for n, row in enumerate(EXAMPLES):
        r = start + n
        for col, idx in text_cols.items():
            C.text_cell(ws.cell(r, col), row[idx])
        for col, idx in input_cols.items():
            C.input_cell(ws.cell(r, col), row[idx], center=col in (8, 9, 10, 20, 21, 22))
        fill_formulas(r)
        ws.row_dimensions[r].height = 56

    last_data = start + len(EXAMPLES)
    for r in range(last_data, last_data + 18):
        for ci in range(1, len(HEADERS) + 1):
            ws.cell(r, ci).border = C.BORDER
        fill_formulas(r)
    last = last_data + 18 - 1

    # dropdowns
    C.add_dropdown(ws, "C", LIB["dropdowns"]["category"], start, last)
    C.add_dropdown(ws, "H", LIB["dropdowns"]["affected"], start, last)
    C.add_dropdown(ws, "Q", LIB["dropdowns"]["control"], start, last)
    C.add_dropdown(ws, "AA", ["Yes", "No"], start, last)
    C.add_dropdown(ws, "AC", ["Yes", "No", "N/A"], start, last)

    for col in ("N", "Y"):
        rng = f"{col}3:{col}{last}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Intolerable"'], fill=PatternFill("solid", fgColor=C.RED), font=Font(name=C.ARIAL, bold=True, color=C.REDDARK)))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"High"'], fill=PatternFill("solid", fgColor=C.ORANGE)))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Medium"'], fill=PatternFill("solid", fgColor=C.YELLOW)))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Low"'], fill=PatternFill("solid", fgColor=C.GREEN)))

    C.add_library_sheet(wb, LIB)
    C.base_font(wb)
    wb.save(path)
    return path


if __name__ == "__main__":
    build("medical_risk_assessment.xlsx")
