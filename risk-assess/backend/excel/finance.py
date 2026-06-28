"""Finance risk assessment workbook — ISO 31000 / COSO ERM / Basel, with risk taxonomy."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from . import common as C
from .library import LIBRARIES

DOMAIN = "finance"
TITLE = "Finance Risk Assessment (ISO 31000 / COSO ERM / Basel taxonomy)"
LIB = LIBRARIES["finance"]


def _scales(wb):
    ws = wb.create_sheet("Scales")
    ws.sheet_view.showGridLines = False
    C.set_widths(ws, {"A": 3, "B": 9, "C": 22, "D": 50, "F": 10, "G": 22, "H": 40, "O": 16, "P": 6})
    C.title(ws["B2"], "Scoring Scales & Risk Acceptability", 14)

    ws["B4"] = "LIKELIHOOD (L)"; C.style_header(ws["B4"]); ws.merge_cells("B4:D4")
    for i, h in enumerate(["Score", "Qualitative", "Annual probability"]):
        C.style_header(ws.cell(5, 2 + i), fill=C.NAVY)
    lk = [(1, "Rare", "< 1%"), (2, "Unlikely", "1-5%"), (3, "Possible", "5-25%"),
          (4, "Likely", "25-50%"), (5, "Almost certain", "> 50%")]
    for j, (s, q, f) in enumerate(lk):
        C.text_cell(ws.cell(6 + j, 2), s, center=True); C.text_cell(ws.cell(6 + j, 3), q); C.text_cell(ws.cell(6 + j, 4), f)

    ws["B12"] = "IMPACT (I) — financial severity"; C.style_header(ws["B12"]); ws.merge_cells("B12:D12")
    for i, h in enumerate(["Score", "Category", "Financial / regulatory impact (illustrative)"]):
        C.style_header(ws.cell(13, 2 + i), fill=C.NAVY)
    im = [(1, "Insignificant", "< $50k; no regulatory attention"),
          (2, "Minor", "$50k-$500k; minor breach"),
          (3, "Moderate", "$500k-$5M; reportable issue"),
          (4, "Major", "$5M-$50M; regulatory action / restatement"),
          (5, "Severe", "> $50M; solvency threat / licence risk")]
    for j, (s, cat, d) in enumerate(im):
        C.text_cell(ws.cell(14 + j, 2), s, center=True); C.text_cell(ws.cell(14 + j, 3), cat); C.text_cell(ws.cell(14 + j, 4), d)
    ws.cell(20, 2, "Expected Annual Loss = Financial Exposure ($) x Annual Probability (%). See Risk Taxonomy sheet.")\
        .font = Font(name=C.ARIAL, italic=True, size=9, color="595959")
    ws.merge_cells("B20:D20")

    thr = C.add_threshold_block(ws, 4, anchor_col=6, score_label="ACCEPTABILITY (Risk Score = L x I, 1-25)")
    C.add_matrix(ws, 12, anchor_col=6, rows_label="Impact", cols_label="Likelihood")
    return thr


HEADERS = [
    ("Risk ID", 9), ("Risk category", 14), ("Risk subcategory", 16), ("Risk event / scenario", 26),
    ("Cause / driver", 22), ("Consequence", 22), ("Risk owner", 12),
    ("L", 5), ("I", 5), ("Inherent Score", 9), ("Inherent Level", 12), ("Acceptability", 14),
    ("Financial Exposure ($)", 15), ("Annual Prob. (%)", 11), ("Expected Annual Loss ($)", 16),
    ("Treatment strategy", 14), ("Control / mitigation", 26), ("Control effectiveness (1-5)", 11),
    ("L'", 5), ("I'", 5), ("Residual Score", 9), ("Residual Level", 12), ("Residual Acceptability", 14),
    ("Key Risk Indicator / trigger", 22), ("KRI vs appetite", 13), ("Status", 10),
]
EXAMPLES = [
    ["FIN-001", "Credit", "Default risk", "Counterparty default in loan portfolio", "Economic downturn raises default rates",
     "Credit losses exceed provisions", "CRO", 3, 4, 5000000, 0.04, "Reduce",
     "Tighten underwriting, collateral & single-name limits", 4, 2, 3, "NPL ratio > 3%", "Within appetite"],
    ["FIN-002", "Market", "Interest-rate risk", "Rate shock erodes bond portfolio", "Rapid central-bank rate moves",
     "Mark-to-market loss on AFS book", "Head of Treasury", 3, 3, 8000000, 0.05, "Reduce",
     "Duration hedging via interest-rate swaps", 4, 2, 3, "DV01 above limit", "Within appetite"],
    ["FIN-003", "Liquidity", "Funding liquidity", "Funding run / deposit outflow", "Loss of market confidence",
     "Inability to meet obligations as due", "Treasurer", 2, 5, 10000000, 0.02, "Reduce",
     "LCR buffer + contingency funding plan", 4, 2, 4, "LCR < 110%", "Breach"],
    ["FIN-004", "Compliance", "AML / KYC", "Regulatory fine for AML control failure", "Gaps in transaction monitoring",
     "Penalty + remediation + reputational harm", "CCO", 2, 4, 3000000, 0.06, "Reduce",
     "Enhanced KYC + automated monitoring + audit", 3, 1, 3, "SAR backlog > 30 days", "Within appetite"],
]


def build(path):
    wb = Workbook()
    wb.remove(wb.active)
    t_int, t_high, t_med = _scales(wb)

    ws = wb.create_sheet("Risk Register", 0)
    ws.sheet_view.showGridLines = False
    ws.cell(1, 1, TITLE).font = Font(name=C.ARIAL, bold=True, color=C.NAVY, size=13)
    ws.row_dimensions[2].height = 42
    for i, (h, w) in enumerate(HEADERS, start=1):
        C.style_header(ws.cell(2, i, h))
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C3"

    def fill_formulas(r):
        ws.cell(r, 10, f'=IF(OR(H{r}="",I{r}=""),"",H{r}*I{r})')          # inherent score
        ws.cell(r, 11, C.level_formula(f"J{r}", t_int, t_high, t_med))
        ws.cell(r, 12, C.accept_formula(f"J{r}", t_int, t_high, t_med))
        ws.cell(r, 15, f'=IF(OR(M{r}="",N{r}=""),"",M{r}*N{r})')          # expected annual loss
        ws.cell(r, 13).number_format = '$#,##0'
        ws.cell(r, 14).number_format = '0.0%'
        ws.cell(r, 15).number_format = '$#,##0'
        ws.cell(r, 21, f'=IF(OR(S{r}="",T{r}=""),"",S{r}*T{r})')          # residual score
        ws.cell(r, 22, C.level_formula(f"U{r}", t_int, t_high, t_med))
        ws.cell(r, 23, C.accept_formula(f"U{r}", t_int, t_high, t_med))
        ws.cell(r, 26, f'=IF(W{r}="","",IF(OR(W{r}="Unacceptable",L{r}="Unacceptable"),"OPEN","Monitor"))')
        for col in (10, 11, 12, 15, 21, 22, 23, 26):
            ws.cell(r, col).alignment = Alignment(horizontal="center", wrap_text=True)
            ws.cell(r, col).border = C.BORDER

    start = 3
    text_cols = {1: 0, 3: 2, 4: 3, 5: 4, 6: 5, 7: 6, 17: 12, 24: 16}
    input_cols = {2: 1, 8: 7, 9: 8, 13: 9, 14: 10, 16: 11, 18: 13, 19: 14, 20: 15, 25: 17}
    for n, row in enumerate(EXAMPLES):
        r = start + n
        for col, idx in text_cols.items():
            C.text_cell(ws.cell(r, col), row[idx])
        for col, idx in input_cols.items():
            C.input_cell(ws.cell(r, col), row[idx], center=col in (8, 9, 18, 19, 20))
        ws.cell(r, 13).number_format = '$#,##0'
        ws.cell(r, 14).number_format = '0.0%'
        fill_formulas(r)
        ws.row_dimensions[r].height = 50

    last_data = start + len(EXAMPLES)
    for r in range(last_data, last_data + 18):
        for ci in range(1, len(HEADERS) + 1):
            ws.cell(r, ci).border = C.BORDER
        fill_formulas(r)
    last = last_data + 18 - 1

    C.add_dropdown(ws, "B", LIB["dropdowns"]["category"], start, last)
    C.add_dropdown(ws, "P", LIB["dropdowns"]["treatment"], start, last)
    C.add_dropdown(ws, "Y", ["Within appetite", "Breach"], start, last)

    for col in ("K", "V"):
        rng = f"{col}3:{col}{last}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Intolerable"'], fill=PatternFill("solid", fgColor=C.RED), font=Font(name=C.ARIAL, bold=True, color=C.REDDARK)))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"High"'], fill=PatternFill("solid", fgColor=C.ORANGE)))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Medium"'], fill=PatternFill("solid", fgColor=C.YELLOW)))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Low"'], fill=PatternFill("solid", fgColor=C.GREEN)))

    C.add_library_sheet(wb, LIB)
    C.base_font(wb)
    wb.save(path)
    return path


if __name__ == "__main__":
    build("finance_risk_assessment.xlsx")
