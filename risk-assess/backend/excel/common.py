"""Shared styling, helpers and risk formulas for all domain Excel templates.

All workbooks follow the same conventions:
  * Blue text  = user input cells (scores, descriptions)
  * Black text = auto-calculated formulas (do not edit)
  * Risk Score = Likelihood/Probability  x  Impact/Severity
  * Risk levels & acceptability looked up against thresholds on the 'Scales' sheet
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ARIAL = "Arial"

# palette
NAVY = "1F3864"
BLUE_HDR = "2E5496"
GREY_HDR = "7F7F7F"
GREEN = "C6EFCE"
YELLOW = "FFEB9C"
ORANGE = "FFD966"
RED = "FFC7CE"
REDDARK = "C00000"
INPUT_BLUE = "0000FF"

_thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

# common 1-5 acceptability thresholds on a 1-25 score (Likelihood x Impact)
THRESHOLDS = {"intolerable": 16, "high": 10, "medium": 5}


def style_header(cell, fill=BLUE_HDR, color="FFFFFF", size=10):
    cell.font = Font(name=ARIAL, bold=True, color=color, size=size)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def title(cell, text, size=15):
    cell.value = text
    cell.font = Font(name=ARIAL, bold=True, color=NAVY, size=size)


def input_cell(cell, value=None, center=True):
    if value is not None:
        cell.value = value
    cell.font = Font(name=ARIAL, color=INPUT_BLUE, size=10)
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="top", wrap_text=True)
    cell.border = BORDER


def text_cell(cell, value=None, bold=False, center=False):
    if value is not None:
        cell.value = value
    cell.font = Font(name=ARIAL, size=10, bold=bold)
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="top", wrap_text=True)
    cell.border = BORDER


def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def base_font(wb):
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None and (c.font is None or c.font.name != ARIAL):
                    c.font = Font(name=ARIAL, size=10)


def level_formula(idx_cell, t_intol, t_high, t_med):
    """Nested IF returning the qualitative risk level for a numeric score cell."""
    return (f'=IF({idx_cell}="","",IF({idx_cell}>={t_intol},"Intolerable",'
            f'IF({idx_cell}>={t_high},"High",IF({idx_cell}>={t_med},"Medium","Low"))))')


def accept_formula(idx_cell, t_intol, t_high, t_med):
    return (f'=IF({idx_cell}="","",IF({idx_cell}>={t_intol},"Unacceptable",'
            f'IF({idx_cell}>={t_high},"Reduce (ALARP)",'
            f'IF({idx_cell}>={t_med},"Acceptable w/ review","Acceptable"))))')


def add_threshold_block(ws, anchor_row, anchor_col=6,
                        levels=(("Intolerable", 16, RED), ("High", 10, ORANGE),
                                ("Medium", 5, YELLOW), ("Low", 1, GREEN)),
                        score_label="Risk Score = Likelihood x Impact (1-25)"):
    """Writes the acceptability legend and returns absolute refs to the
    intolerable/high/medium threshold value cells for use in formulas."""
    c = anchor_col
    cl = get_column_letter(c)
    ws.cell(anchor_row, c, score_label)
    style_header(ws.cell(anchor_row, c), fill=NAVY)
    ws.merge_cells(start_row=anchor_row, start_column=c, end_row=anchor_row, end_column=c + 2)
    for i, h in enumerate(["Score >=", "Risk Level", "Action / Acceptability"]):
        style_header(ws.cell(anchor_row + 1, c + i), fill=NAVY)
    actions = {
        "Intolerable": "Unacceptable - must reduce risk before proceeding",
        "High": "Undesirable - reduce as far as possible (ALARP); justify residual",
        "Medium": "Acceptable with review / documented controls",
        "Low": "Broadly acceptable - no further action required",
    }
    for j, (lvl, thr, fill) in enumerate(levels):
        rr = anchor_row + 2 + j
        text_cell(ws.cell(rr, c), thr, center=True)
        text_cell(ws.cell(rr, c + 1), lvl, bold=True)
        text_cell(ws.cell(rr, c + 2), actions[lvl])
        for cc in range(c, c + 3):
            ws.cell(rr, cc).fill = PatternFill("solid", fgColor=fill)
    # machine-referenced threshold cells — placed in a fixed far-right area
    # (columns O/P) so the risk matrix or legend can never overwrite them.
    hc = 15  # column O (label), P (value)
    ws.cell(4, hc, "Thresholds used by formulas (do not move):").font = Font(name=ARIAL, italic=True, size=9, color="595959")
    ws.cell(5, hc, "Intolerable >=").font = Font(name=ARIAL, size=9)
    ws.cell(5, hc + 1, THRESHOLDS["intolerable"]).font = Font(name=ARIAL, size=9, bold=True)
    ws.cell(6, hc, "High >=").font = Font(name=ARIAL, size=9)
    ws.cell(6, hc + 1, THRESHOLDS["high"]).font = Font(name=ARIAL, size=9, bold=True)
    ws.cell(7, hc, "Medium >=").font = Font(name=ARIAL, size=9)
    ws.cell(7, hc + 1, THRESHOLDS["medium"]).font = Font(name=ARIAL, size=9, bold=True)
    sheet = ws.title
    tcol = get_column_letter(hc + 1)  # P
    return (f"'{sheet}'!${tcol}$5", f"'{sheet}'!${tcol}$6", f"'{sheet}'!${tcol}$7")


def add_dropdown(ws, col_letter, options, first_row, last_row, prompt=None):
    """Attach a list data-validation (dropdown) to a column range."""
    # Excel inline list limit ~255 chars; our option sets are short.
    formula = '"' + ",".join(o.replace(",", " /") for o in options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    if prompt:
        dv.prompt = prompt
        dv.promptTitle = "Select"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first_row}:{col_letter}{last_row}")
    return dv


def add_library_sheet(wb, spec):
    """Builds a reference 'mapping' sheet from a library spec dict."""
    ws = wb.create_sheet(spec["sheet"])
    ws.sheet_view.showGridLines = False
    title(ws["B2"], spec["title"], 13)
    cols = spec["columns"]
    widths = [22, 30, 34, 30, 26, 22]
    for i, h in enumerate(cols):
        ws.column_dimensions[get_column_letter(2 + i)].width = widths[i] if i < len(widths) else 22
        style_header(ws.cell(4, 2 + i, h), fill=NAVY)
    ws.column_dimensions["A"].width = 3
    for r, row in enumerate(spec["rows"], start=5):
        for i, val in enumerate(row):
            cell = ws.cell(r, 2 + i, val)
            cell.font = Font(name=ARIAL, size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top",
                                       horizontal="center" if isinstance(val, int) else "left")
            cell.border = BORDER
        ws.row_dimensions[r].height = 30
        if r % 2 == 0:
            for i in range(len(row)):
                ws.cell(r, 2 + i).fill = PatternFill("solid", fgColor="F2F5FB")
    ws.freeze_panes = "A5"
    return ws


def add_matrix(ws, anchor_row, anchor_col=2, rows_label="Impact", cols_label="Likelihood"):
    """5x5 colour-coded score matrix (value = row index x col index)."""
    c = anchor_col
    ws.cell(anchor_row, c, f"RISK MATRIX (cell = {rows_label[:4]} x {cols_label[:4]})").font = \
        Font(name=ARIAL, bold=True, color=NAVY, size=11)
    hc = ws.cell(anchor_row + 1, c, f"{rows_label}\\{cols_label}")
    style_header(hc, fill=NAVY)
    for p in range(1, 6):
        style_header(ws.cell(anchor_row + 1, c + p, p), fill=NAVY)
    for s in range(5, 0, -1):
        rr = anchor_row + 1 + (6 - s)
        style_header(ws.cell(rr, c, s), fill=NAVY)
        for p in range(1, 6):
            val = s * p
            cell = ws.cell(rr, c + p, val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER
            cell.font = Font(name=ARIAL, size=10, bold=True)
            fill = RED if val >= 16 else ORANGE if val >= 10 else YELLOW if val >= 5 else GREEN
            cell.fill = PatternFill("solid", fgColor=fill)
