"""Technology risk assessment workbook — Reputation / Financial / IP loss, with STRIDE threat catalog."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from . import common as C
from .library import LIBRARIES

DOMAIN = "technology"
TITLE = "Technology Risk Assessment (Reputation / Financial / IP loss) — NIST / ISO 27005 / FAIR"
LIB = LIBRARIES["technology"]


def _scales(wb):
    ws = wb.create_sheet("Scales")
    ws.sheet_view.showGridLines = False
    C.set_widths(ws, {"A": 3, "B": 9, "C": 20, "D": 28, "E": 28, "G": 10, "H": 22, "I": 40, "O": 16, "P": 6})
    C.title(ws["B2"], "Scoring Scales & Risk Acceptability", 14)

    ws["B4"] = "LIKELIHOOD (L)"; C.style_header(ws["B4"]); ws.merge_cells("B4:E4")
    for i, h in enumerate(["Score", "Qualitative", "Meaning"]):
        C.style_header(ws.cell(5, 2 + i), fill=C.NAVY)
    C.style_header(ws.cell(5, 4), fill=C.NAVY)
    ws.merge_cells("D5:E5")
    lk = [(1, "Rare", "Not expected in 5 yrs"), (2, "Unlikely", "Once in 2-5 yrs"),
          (3, "Possible", "Once a year"), (4, "Likely", "Several times a year"), (5, "Almost certain", "Monthly or more")]
    for j, (s, q, m) in enumerate(lk):
        C.text_cell(ws.cell(6 + j, 2), s, center=True); C.text_cell(ws.cell(6 + j, 3), q)
        ws.merge_cells(start_row=6 + j, start_column=4, end_row=6 + j, end_column=5)
        C.text_cell(ws.cell(6 + j, 4), m)

    ws["B12"] = "IMPACT (1-5) per loss dimension"; C.style_header(ws["B12"]); ws.merge_cells("B12:E12")
    for i, h in enumerate(["Score", "Reputation loss", "Financial loss", "IP loss"]):
        C.style_header(ws.cell(13, 2 + i), fill=C.NAVY)
    im = [(1, "No external notice", "< $50k", "No proprietary exposure"),
          (2, "Local complaints", "$50k-$500k", "Minor / non-core IP exposed"),
          (3, "Negative press / social", "$500k-$5M", "Some trade secrets at risk"),
          (4, "National media / churn", "$5M-$50M", "Core IP / source code exposed"),
          (5, "Brand crisis, lasting", "> $50M", "Critical IP lost / patents void")]
    for j, (s, rep, fin, ip) in enumerate(im):
        C.text_cell(ws.cell(14 + j, 2), s, center=True)
        C.text_cell(ws.cell(14 + j, 3), rep); C.text_cell(ws.cell(14 + j, 4), fin); C.text_cell(ws.cell(14 + j, 5), ip)
    ws.cell(20, 2, "Aggregate Impact = MAX(Reputation, Financial, IP). Risk Score = Likelihood x Aggregate Impact.")\
        .font = Font(name=C.ARIAL, italic=True, size=9, color="595959")
    ws.merge_cells("B20:E20")

    thr = C.add_threshold_block(ws, 4, anchor_col=7, score_label="ACCEPTABILITY (Score = L x AggImpact, 1-25)")
    C.add_matrix(ws, 12, anchor_col=7, rows_label="Impact", cols_label="Likelihood")
    return thr


HEADERS = [
    ("Risk ID", 10), ("Asset / system", 18), ("STRIDE category", 16), ("Threat / risk event", 24),
    ("Threat source / actor", 16), ("Vulnerability / cause", 22), ("Consequence", 22), ("Risk owner", 12), ("L", 5),
    ("Reputation Impact", 11), ("Financial Impact", 11), ("IP Loss Impact", 11), ("Aggregate Impact", 11),
    ("Inherent Score", 9), ("Inherent Level", 12), ("Acceptability", 13),
    ("Existing controls", 22), ("Additional control / mitigation", 26), ("Control type", 16), ("CVE / SBOM / ref", 16),
    ("L'", 5), ("Rep'", 6), ("Fin'", 6), ("IP'", 6), ("Agg Impact'", 10), ("Residual Score", 9),
    ("Residual Level", 12), ("Residual Acceptability", 13), ("Primary loss type", 13), ("Status", 10),
]
EXAMPLES = [
    ["TECH-001", "Customer database / web app", "Information disclosure", "Customer data breach (PII)",
     "External attacker", "Exploited web vulnerability; no WAF", "Regulatory fines + churn + brand damage", "CISO",
     3, 5, 4, 2, "TLS in transit; basic logging", "WAF + encryption at rest + IR plan + quarterly pen-test",
     "Protective measure", "CVE-2024-XXXX; SBOM-3.1", 2, 3, 3, 2],
    ["TECH-002", "Source-code repository", "IP theft", "Source-code / trade-secret theft",
     "Malicious insider", "Weak access controls; no DLP", "Loss of competitive advantage; IP erosion", "CTO",
     2, 3, 3, 5, "SSO; repo permissions", "DLP + least-privilege + repo access monitoring + NDAs",
     "Detective control", "SBOM-3.1", 1, 2, 2, 3],
    ["TECH-003", "Production SaaS platform", "Denial of service", "Critical SaaS outage",
     "External / botnet", "Single-region dependency; no failover", "SLA penalties + revenue loss + reputation", "VP Engineering",
     3, 4, 4, 1, "CDN; autoscaling", "Multi-region HA + DR runbook + chaos testing",
     "Inherently secure design", "DR-PLAN-2", 2, 3, 3, 1],
    ["TECH-004", "Software supply chain", "Supply chain", "Open-source licence / IP infringement",
     "Third-party dependency", "Unvetted copyleft dependency", "Forced disclosure of proprietary code; litigation", "Legal + Eng",
     3, 2, 3, 4, "Manual review", "SBOM + automated licence scanning in CI + approval gate",
     "Procedural control", "SBOM-3.1; SCA-7", 1, 1, 2, 2],
]


def build(path):
    wb = Workbook()
    wb.remove(wb.active)
    t_int, t_high, t_med = _scales(wb)

    ws = wb.create_sheet("Risk Register", 0)
    ws.sheet_view.showGridLines = False
    ws.cell(1, 1, TITLE).font = Font(name=C.ARIAL, bold=True, color=C.NAVY, size=12)
    ws.row_dimensions[2].height = 44
    for i, (h, w) in enumerate(HEADERS, start=1):
        C.style_header(ws.cell(2, i, h))
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C3"

    def fill_formulas(r):
        ws.cell(r, 13, f'=IF(COUNT(J{r}:L{r})=0,"",MAX(J{r}:L{r}))')        # aggregate impact
        ws.cell(r, 14, f'=IF(OR(I{r}="",M{r}=""),"",I{r}*M{r})')            # inherent score
        ws.cell(r, 15, C.level_formula(f"N{r}", t_int, t_high, t_med))
        ws.cell(r, 16, C.accept_formula(f"N{r}", t_int, t_high, t_med))
        ws.cell(r, 25, f'=IF(COUNT(V{r}:X{r})=0,"",MAX(V{r}:X{r}))')        # residual aggregate
        ws.cell(r, 26, f'=IF(OR(U{r}="",Y{r}=""),"",U{r}*Y{r})')            # residual score
        ws.cell(r, 27, C.level_formula(f"Z{r}", t_int, t_high, t_med))
        ws.cell(r, 28, C.accept_formula(f"Z{r}", t_int, t_high, t_med))
        ws.cell(r, 29, f'=IF(M{r}="","",IF(AND(J{r}>=K{r},J{r}>=L{r}),"Reputation",IF(AND(K{r}>=J{r},K{r}>=L{r}),"Financial","IP")))')
        ws.cell(r, 30, f'=IF(AB{r}="","",IF(OR(AB{r}="Unacceptable",P{r}="Unacceptable"),"OPEN","Monitor"))')
        for col in (13, 14, 15, 16, 25, 26, 27, 28, 29, 30):
            ws.cell(r, col).alignment = Alignment(horizontal="center", wrap_text=True)
            ws.cell(r, col).border = C.BORDER

    start = 3
    text_cols = {1: 0, 2: 1, 4: 3, 5: 4, 6: 5, 7: 6, 8: 7, 17: 12, 18: 13, 20: 15}
    input_cols = {3: 2, 9: 8, 10: 9, 11: 10, 12: 11, 19: 14, 21: 16, 22: 17, 23: 18, 24: 19}
    for n, row in enumerate(EXAMPLES):
        r = start + n
        for col, idx in text_cols.items():
            C.text_cell(ws.cell(r, col), row[idx])
        for col, idx in input_cols.items():
            C.input_cell(ws.cell(r, col), row[idx], center=col in (9, 10, 11, 12, 21, 22, 23, 24))
        fill_formulas(r)
        ws.row_dimensions[r].height = 56

    last_data = start + len(EXAMPLES)
    for r in range(last_data, last_data + 18):
        for ci in range(1, len(HEADERS) + 1):
            ws.cell(r, ci).border = C.BORDER
        fill_formulas(r)
    last = last_data + 18 - 1

    C.add_dropdown(ws, "C", LIB["dropdowns"]["stride"], start, last)
    C.add_dropdown(ws, "S", LIB["dropdowns"]["control"], start, last)

    for col in ("O", "AA"):
        rng = f"{col}3:{col}{last}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Intolerable"'], fill=PatternFill("solid", fgColor=C.RED), font=Font(name=C.ARIAL, bold=True, color=C.REDDARK)))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"High"'], fill=PatternFill("solid", fgColor=C.ORANGE)))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Medium"'], fill=PatternFill("solid", fgColor=C.YELLOW)))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Low"'], fill=PatternFill("solid", fgColor=C.GREEN)))

    C.add_library_sheet(wb, LIB)
    C.base_font(wb)
    wb.save(path)
    return path


if __name__ == "__main__":
    build("technology_risk_assessment.xlsx")
