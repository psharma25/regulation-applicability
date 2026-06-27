"""Build a prioritised, actionable implementation roadmap from a set of
applicable regulations, and render it as an Excel workbook (openpyxl).

When multiple regulations are selected, overlapping controls are de-duplicated
(implement-once / comply-many) and tagged with every regulation they satisfy.
"""
import io
from typing import List, Dict
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .. import models

PHASE_NAME = {1: "Foundation (M1-3)", 2: "Core Controls (M3-6)",
              3: "Hardening (M6-10)", 4: "Scale & Audit (M10-15)",
              5: "Advanced (M15+)"}
PRIO_FILL = {"Critical": "C0504D", "High": "E8A020", "Moderate": "0EA5C9", "Low": "94A3B8"}


def build(db: Session, reg_ids: List[str]) -> List[Dict]:
    """Return de-duplicated controls applicable to the selected regulations,
    each tagged with the regulations it satisfies, ordered for execution."""
    reg_ids = set(reg_ids)
    rows = []
    for c in db.query(models.Control).all():
        covered = sorted(set(c.regulations or []) & reg_ids)
        if not covered:
            continue
        rows.append({
            "id": c.id, "domain": c.domain, "title": c.title,
            "explanation": c.explanation, "actions": c.actions or [],
            "priority": c.priority, "severity": c.severity, "score": c.score,
            "phase": c.phase, "effort": c.effort, "regulations": covered,
        })
    order = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
    rows.sort(key=lambda r: (r["phase"], order.get(r["priority"], 9), -r["score"]))
    for i, r in enumerate(rows, 1):
        r["seq"] = i
    return rows


def to_excel(db: Session, reg_ids: List[str], profile_name: str = "My platform") -> bytes:
    rows = build(db, reg_ids)
    regs = {r.id: r for r in db.query(models.Regulation).all()}
    wb = Workbook()

    # ---- Sheet 1: Roadmap ------------------------------------------------
    ws = wb.active
    ws.title = "Roadmap"
    headers = ["#", "Phase", "Priority", "Severity", "Score", "Domain", "Control",
               "Why it matters", "Action steps", "Applicable regulations",
               "Effort", "Owner", "Status", "Target date", "Evidence"]
    ws.merge_cells("A1:O1")
    t = ws.cell(1, 1, f"Implementation Roadmap \u2014 {profile_name}")
    t.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="0D1B3E")
    t.alignment = Alignment("left", "center")
    ws.row_dimensions[1].height = 26
    for c, h in enumerate(headers, 1):
        cell = ws.cell(2, c, h)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="152448")
        cell.alignment = Alignment("center", "center", wrap_text=True)
    thin = Side(style="thin", color="D7DEEA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    r = 3
    for row in rows:
        actions = "\n".join(f"{i+1}. {a}" for i, a in enumerate(row["actions"]))
        vals = [row["seq"], PHASE_NAME.get(row["phase"], row["phase"]), row["priority"],
                row["severity"], row["score"], row["domain"], row["title"],
                row["explanation"], actions, ", ".join(row["regulations"]),
                row["effort"], "", "Not started", "", ""]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment("left", "top", wrap_text=True)
            cell.border = border
            if c == 3:  # priority colour
                cell.fill = PatternFill("solid", fgColor=PRIO_FILL.get(
                    {"Critical": "Critical", "High": "High", "Moderate": "Moderate",
                     "Low": "Low"}.get(row["severity"], "Low"), "FFFFFF"))
                cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                cell.alignment = Alignment("center", "center")
        ws.row_dimensions[r].height = 64
        r += 1
    widths = [4, 16, 10, 10, 7, 14, 26, 40, 46, 22, 7, 12, 12, 12, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"

    # ---- Sheet 2: Regulations covered -----------------------------------
    ws2 = wb.create_sheet("Regulations")
    h2 = ["ID", "Regulation", "Body", "Domain", "Jurisdictions", "Emerging",
          "Official source", "Summary"]
    for c, h in enumerate(h2, 1):
        cell = ws2.cell(1, c, h)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="152448")
        cell.alignment = Alignment("center", "center", wrap_text=True)
    rr = 2
    for rid in sorted(reg_ids):
        reg = regs.get(rid)
        if not reg:
            continue
        vals = [reg.id, reg.name, reg.body, reg.domain,
                ", ".join(reg.jurisdictions or []), "Yes" if reg.emerging else "No",
                reg.url, reg.summary]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(rr, c, v)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment("left", "top", wrap_text=True)
        rr += 1
    for i, w in enumerate([12, 30, 14, 14, 22, 9, 42, 60], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
